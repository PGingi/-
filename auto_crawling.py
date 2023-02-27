from pytube import YouTube
from openpyxl import Workbook
from openpyxl import load_workbook
#import ssl
#ssl._create_default_https_context = ssl._create_stdlib_context

load_wb = load_workbook('D:\화재영상\화재영상url.xlsx', data_only=True)
load_ws = load_wb['화재영상']

DownloadFolder = "D:\화재영상"
urls = []
for i in range(30):
    vlaues = load_ws.cell(1+i,1).value
    urls.append(vlaues)
print(urls)

for url in urls:
    video = YouTube(url)    
    #print("title : ", yt.title)
    #print("length : ", yt.length)
    #print("author : ", yt.author)
    #print("publish_date : ", yt.publish_date)
    #print("views : ", yt.views)
    #print("keywords : ", yt.keywords)
    #print("description : ", yt.description)
    clip = video.streams.get_highest_resolution()
    clip.download(DownloadFolder)